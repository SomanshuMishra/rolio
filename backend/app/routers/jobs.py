from fastapi import APIRouter, Depends, HTTPException, status
from starlette.responses import FileResponse
from sqlalchemy.orm import Session
from datetime import datetime
import logging
import time
import uuid
import json
import io
import tempfile

from ..database import get_db
from ..models import User, Resume, UserPreferences, APIKey, UserJobMatch, UserJobAction, Job
from ..schemas.job import (
    JobSearchRequest,
    JobSearchResponse,
    JobMatchResult,
    JobMatches,
    JobMatchWithAction,
)
from ..utils.deps import get_current_user
from ..utils.security import decrypt_api_key
from ..services.jsearch_client import JSearchClient, cache_jobs, get_non_expired_jobs, cleanup_expired_jobs
from ..services.job_matcher import JobMatcher
from ..services.web_job_searcher import GeminiWebJobSearcher
from ..services.background_jobs import start_background_search
from ..models import JobSearch

router = APIRouter(prefix="/jobs", tags=["jobs"])
logger = logging.getLogger(__name__)

# Skills database for filtering
AVAILABLE_SKILLS = [
    "Python", "Java", "JavaScript", "TypeScript", "C++", "C#", "Go", "Rust", "Ruby", "PHP",
    "Swift", "Kotlin", "Scala", "Groovy", "R", "MATLAB", "SQL", "Bash", "Shell",
    "Django", "Flask", "FastAPI", "Spring", "Spring Boot", "Express", "React", "Vue", "Angular",
    "Next.js", "Svelte", "Rails", "Laravel", "ASP.NET", "Blazor",
    "PostgreSQL", "MySQL", "MongoDB", "Redis", "Cassandra", "DynamoDB", "Elasticsearch",
    "Oracle", "SQL Server", "MariaDB", "Firebase", "GraphQL",
    "AWS", "Azure", "GCP", "Docker", "Kubernetes", "Jenkins", "GitLab", "GitHub", "CircleCI",
    "Travis CI", "Terraform", "Ansible", "CloudFormation", "EC2", "S3", "Lambda",
    "Git", "REST API", "Microservices", "Celery", "RabbitMQ", "Kafka",
    "NumPy", "Pandas", "Scikit-learn", "TensorFlow", "PyTorch", "OpenCV",
    "HTML", "CSS", "Sass", "Bootstrap", "Tailwind", "Material UI",
    "Machine Learning", "Deep Learning", "NLP", "Data Science",
    "Big Data", "Spark", "Hadoop", "ETL", "CI/CD", "Agile", "Scrum",
]


@router.get("/skills")
def get_available_skills():
    """Get list of available skills for filtering."""
    return {"skills": sorted(AVAILABLE_SKILLS)}


@router.post("/search", response_model=JobSearchResponse, status_code=status.HTTP_200_OK)
async def search_and_match_jobs(
    request: JobSearchRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Trigger AI-powered job search and matching."""
    start_time = time.time()

    logger.info(f"\n{'='*80}")
    logger.info(f"SEARCH REQUEST from user {current_user.id} ({current_user.email})")
    logger.info(f"{'='*80}\n")

    try:
        # Check user has resume
        logger.info(f"1. Checking for resume...")
        resume = db.query(Resume).filter(
            Resume.user_id == current_user.id,
            Resume.is_active == True,
        ).first()

        if not resume:
            logger.error(f"   ❌ No resume found for user {current_user.id}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No resume uploaded. Please upload a resume first.",
            )
        logger.info(f"   ✓ Resume found: {resume.filename}")

        # Get user preferences
        logger.info(f"2. Checking for preferences...")
        preferences = db.query(UserPreferences).filter(
            UserPreferences.user_id == current_user.id,
        ).first()

        if not preferences:
            logger.error(f"   ❌ No preferences found for user {current_user.id}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No preferences set. Please set your job preferences first.",
            )
        logger.info(f"   ✓ Preferences found:")
        logger.info(f"     - Roles: {preferences.preferred_roles}")
        logger.info(f"     - Locations: {preferences.preferred_locations}")
        logger.info(f"     - Salary: {preferences.salary_min} - {preferences.salary_max}")
        logger.info(f"     - Remote: {preferences.remote_preference}")

        # Get user's default API key
        logger.info(f"3. Checking for default API key...")
        # Expunge any stale session objects
        db.expunge_all()

        api_key_record = db.query(APIKey).filter(
            APIKey.user_id == current_user.id,
            APIKey.is_default == True,
        ).first()

        if not api_key_record:
            logger.warning(f"   ⚠ No default API key found, trying any available key...")
            api_key_record = db.query(APIKey).filter(
                APIKey.user_id == current_user.id,
            ).order_by(APIKey.created_at.desc()).first()

        if not api_key_record:
            logger.error(f"   ❌ No API key found for user {current_user.id}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No AI provider API key configured. Please add your OpenAI or Claude API key.",
            )
        logger.info(f"   ✓ API key found: {api_key_record.provider} (default={api_key_record.is_default})")

        # Decrypt API key
        api_key = decrypt_api_key(api_key_record.encrypted_key)
        if not api_key:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to decrypt API key",
            )

        # Clean up expired cache
        cleanup_expired_jobs(db)

        # Get or fetch jobs from multiple sources
        from ..services.job_discovery import (
            fetch_remoteok_jobs,
            fetch_jsearch_smart_queries,
            fetch_gemini_web_jobs,
            merge_and_dedupe_jobs,
        )

        jobs = []
        if not request.force_refresh:
            # Try to get from cache first
            cached_jobs = get_non_expired_jobs(db, limit=request.limit * 5)
            jobs = [
                {
                    "jsearch_id": j.jsearch_id,
                    "title": j.title,
                    "company": j.company,
                    "location": j.location,
                    "is_remote": j.is_remote,
                    "salary_min": j.salary_min,
                    "salary_max": j.salary_max,
                    "description": j.description,
                    "requirements": [r.strip() for r in j.requirements.split(",") if r.strip()] if j.requirements else [],
                    "apply_url": j.apply_url,
                    "posted_at": j.posted_at,
                }
                for j in cached_jobs
            ]

        # If not enough cached jobs, fetch from multiple sources
        logger.info(f"4. Cached jobs: {len(jobs)}, need: {request.limit * 2}")
        if len(jobs) < request.limit * 2 or request.force_refresh:
            try:
                # Parse preferences and resume for job discovery
                resume_data = json.loads(resume.parsed_data) if isinstance(resume.parsed_data, str) else resume.parsed_data
                resume_text = resume.raw_text or json.dumps(resume_data)

                prefs_dict = {
                    "preferred_roles": [r.strip() for r in preferences.preferred_roles.split(",") if r.strip()] if preferences.preferred_roles else [],
                    "preferred_locations": [l.strip() for l in preferences.preferred_locations.split(",") if l.strip()] if preferences.preferred_locations else [],
                    "salary_min": preferences.salary_min,
                    "salary_max": preferences.salary_max,
                    "remote_preference": preferences.remote_preference,
                }

                logger.info(f"\n5. Multi-source job discovery:")

                # 1. Fetch from RemoteOK (always)
                logger.info(f"   a) Fetching from RemoteOK...")
                remoteok_jobs = await fetch_remoteok_jobs(limit=60)

                # 2. Fetch from JSearch with smart AI-generated queries (always)
                logger.info(f"   b) Fetching from JSearch with smart queries...")
                jsearch_jobs = await fetch_jsearch_smart_queries(
                    ai_provider=api_key_record.provider,
                    api_key=api_key,
                    resume_text=resume_text,
                    preferences=prefs_dict,
                    limit=60,
                )

                # 3. Fetch from Gemini web search (only if Gemini is selected)
                gemini_jobs = []
                if api_key_record.provider == "google":
                    logger.info(f"   c) Fetching from Gemini web search...")
                    gemini_jobs = await fetch_gemini_web_jobs(
                        api_key=api_key,
                        resume_text=resume_text,
                        preferences=prefs_dict,
                        limit=60,
                    )

                # Merge and dedupe all sources
                logger.info(f"   d) Merging and deduplicating...")
                merged_jobs = merge_and_dedupe_jobs(remoteok_jobs, jsearch_jobs, gemini_jobs)

                if merged_jobs:
                    await cache_jobs(db, merged_jobs)
                    jobs.extend(merged_jobs)
                    logger.info(f"   ✓ Multi-source job discovery returned {len(merged_jobs)} unique jobs")
                    logger.info(f"   Total jobs available: {len(jobs)}")
                else:
                    logger.warning(f"   ⚠ No jobs found from any source!")
                    if not jobs:
                        raise HTTPException(
                            status_code=status.HTTP_404_NOT_FOUND,
                            detail="No jobs found. Try adjusting your preferences.",
                        )

            except HTTPException:
                raise
            except Exception as e:
                logger.error(f"\n   ❌ Error in multi-source job discovery: {str(e)}")
                logger.error(f"      Exception type: {type(e).__name__}")
                import traceback
                logger.error(f"      Traceback: {traceback.format_exc()}")
                if not jobs:
                    raise HTTPException(
                        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                        detail=f"Failed to fetch jobs: {str(e)}",
                    )

        if not jobs:
            logger.error(f"\n   ❌ No jobs available for matching!")
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No jobs found. Try adjusting your preferences.",
            )

        # Limit jobs to match to avoid timeout (match top 30 only)
        jobs_to_match = jobs[:min(30, len(jobs))]

        logger.info(f"\n6. Starting job matching with AI provider: {api_key_record.provider}")
        logger.info(f"   Total jobs to match: {len(jobs_to_match)} (limited to 30 for performance)")

        # Match jobs using AI
        matcher = JobMatcher(api_key_record.provider, api_key)

        # Parse comma-separated strings to lists
        preferred_roles = [r.strip() for r in preferences.preferred_roles.split(",") if r.strip()] if preferences.preferred_roles else []
        preferred_locations = [l.strip() for l in preferences.preferred_locations.split(",") if l.strip()] if preferences.preferred_locations else []

        preferences_dict = {
            "preferred_roles": preferred_roles,
            "preferred_locations": preferred_locations,
            "salary_min": preferences.salary_min,
            "salary_max": preferences.salary_max,
            "remote_preference": preferences.remote_preference,
        }

        # Deserialize parsed_data from JSON string if needed (SQLite compatibility)
        parsed_resume = json.loads(resume.parsed_data) if isinstance(resume.parsed_data, str) else resume.parsed_data

        logger.info(f"   Resume: {parsed_resume.get('name', 'Unknown')} with {len(parsed_resume.get('skills', []))} skills")

        try:
            matches = await matcher.match_jobs(
                parsed_resume,
                jobs_to_match,
                preferences_dict,
                required_skills=request.required_skills if request.required_skills else None,
            )
            logger.info(f"   ✓ Matching complete! Found {len(matches)} matches")
        except Exception as e:
            logger.error(f"\n   ❌ Error during job matching: {str(e)}")
            logger.error(f"      Exception type: {type(e).__name__}")
            import traceback
            logger.error(f"      Traceback: {traceback.format_exc()}")
            raise

        # Store matches in database
        for match in matches[:request.limit]:
            # Get or create job record
            job = db.query(Job).filter(Job.jsearch_id == match["job"]["jsearch_id"]).first()

            if not job:
                # Convert requirements list to comma-separated string
                requirements = match["job"].get("requirements", [])
                requirements_str = ",".join(str(r).strip() for r in requirements if r) if requirements else ""

                job = Job(
                    jsearch_id=match["job"]["jsearch_id"],
                    title=match["job"].get("title"),
                    company=match["job"].get("company"),
                    location=match["job"].get("location"),
                    is_remote=match["job"].get("is_remote", False),
                    salary_min=match["job"].get("salary_min"),
                    salary_max=match["job"].get("salary_max"),
                    description=match["job"].get("description"),
                    requirements=requirements_str,
                    apply_url=match["job"].get("apply_url"),
                    posted_at=match["job"].get("posted_at"),
                    source="jsearch",
                )
                db.add(job)
                db.flush()

            # Create or update match record
            existing_match = db.query(UserJobMatch).filter(
                UserJobMatch.user_id == current_user.id,
                UserJobMatch.job_id == job.id,
            ).first()

            match_reasons_dict = {"reasons": match["match_reasons"]}
            match_reasons_json = json.dumps(match_reasons_dict)

            if existing_match:
                existing_match.match_score = match["match_score"]
                existing_match.match_reasons = match_reasons_json
                existing_match.embedding_score = match["embedding_score"]
            else:
                user_job_match = UserJobMatch(
                    user_id=current_user.id,
                    job_id=job.id,
                    match_score=match["match_score"],
                    match_reasons=match_reasons_json,
                    embedding_score=match["embedding_score"],
                    salary_match=match.get("salary_match", True),
                    location_match=match.get("location_match", True),
                )
                db.add(user_job_match)

        db.commit()

        # Convert matches to response format
        response_jobs = []
        for match in matches[:request.limit]:
            response_jobs.append(
                JobMatchResult(
                    id=uuid.uuid4(),
                    jsearch_id=match["job"]["jsearch_id"],
                    title=match["job"].get("title"),
                    company=match["job"].get("company"),
                    location=match["job"].get("location"),
                    is_remote=match["job"].get("is_remote", False),
                    salary_min=match["job"].get("salary_min"),
                    salary_max=match["job"].get("salary_max"),
                    description=match["job"].get("description"),
                    apply_url=match["job"].get("apply_url"),
                    posted_at=match["job"].get("posted_at"),
                    match_score=match["match_score"],
                    match_reasons=match["match_reasons"],
                    matching_skills=match.get("matching_skills", []),
                    salary_match=match.get("salary_match", True),
                    location_match=match.get("location_match", True),
                )
            )

        processing_time = int((time.time() - start_time) * 1000)

        return JobSearchResponse(
            search_id=str(uuid.uuid4()),
            total_matches=len(matches),
            matches_returned=len(response_jobs),
            processing_time_ms=processing_time,
            jobs=response_jobs,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in job search: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Job search failed: {str(e)}",
        )


@router.get("/matches", response_model=JobMatches)
def get_job_matches(
    limit: int = 20,
    offset: int = 0,
    min_score: int = 60,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get user's saved job matches."""
    logger.info(f"\n{'='*80}")
    logger.info(f"MATCHES REQUEST from user {current_user.id} ({current_user.email})")
    logger.info(f"Limit: {limit}, Offset: {offset}, Min Score: {min_score}")
    logger.info(f"{'='*80}\n")

    try:
        # Query matches
        query = db.query(UserJobMatch).filter(
            UserJobMatch.user_id == current_user.id,
            UserJobMatch.match_score >= min_score,
        )

        total = query.count()
        logger.info(f"Total matches in DB: {total}")

        matches = (
            query.order_by(UserJobMatch.match_score.desc())
            .offset(offset)
            .limit(limit)
            .all()
        )

        result_matches = []
        for match in matches:
            # Get job details
            job = db.query(Job).filter(Job.id == match.job_id).first()
            if not job:
                continue

            # Get user action if exists
            action = db.query(UserJobAction).filter(
                UserJobAction.user_id == current_user.id,
                UserJobAction.job_id == match.job_id,
            ).first()

            # Deserialize match_reasons from JSON string if needed (SQLite compatibility)
            match_reasons_data = match.match_reasons
            if isinstance(match_reasons_data, str):
                match_reasons_data = json.loads(match_reasons_data) if match_reasons_data else {}
            match_reasons_list = match_reasons_data.get("reasons", []) if match_reasons_data else []

            result_matches.append(
                JobMatchWithAction(
                    match_id=match.id,
                    job=JobMatchResult(
                        id=job.id,
                        jsearch_id=job.jsearch_id,
                        title=job.title,
                        company=job.company,
                        location=job.location,
                        is_remote=job.is_remote,
                        salary_min=job.salary_min,
                        salary_max=job.salary_max,
                        description=job.description,
                        apply_url=job.apply_url,
                        posted_at=job.posted_at,
                        match_score=match.match_score,
                        match_reasons=match_reasons_list,
                        salary_match=match.salary_match,
                        location_match=match.location_match,
                    ),
                    match_score=match.match_score,
                    match_reasons=match_reasons_list,
                    user_action=action.action if action else None,
                    created_at=match.created_at,
                )
            )

        return JobMatches(total=total, limit=limit, offset=offset, matches=result_matches)

    except Exception as e:
        logger.error(f"Error fetching matches: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch matches",
        )


@router.post("/jobs/{job_id}/save")
def save_job(
    job_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Save a job to user's saved list."""
    try:
        # Get job
        job = db.query(Job).filter(Job.jsearch_id == job_id).first()

        if not job:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Job not found",
            )

        # Check if already saved
        existing_action = db.query(UserJobAction).filter(
            UserJobAction.user_id == current_user.id,
            UserJobAction.job_id == job.id,
            UserJobAction.action == "saved",
        ).first()

        if existing_action:
            return {"message": "Job already saved", "action": "saved", "job_id": job.id}

        # Create action record
        action = UserJobAction(
            user_id=current_user.id,
            job_id=job.id,
            action="saved",
        )
        db.add(action)
        db.commit()

        logger.info(f"User {current_user.id} saved job {job_id}")

        return {"message": "Job saved", "action": "saved", "job_id": job.id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error saving job: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to save job",
        )


@router.post("/jobs/{job_id}/dismiss")
def dismiss_job(
    job_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Dismiss a job (don't show similar matches)."""
    try:
        # Get job
        job = db.query(Job).filter(Job.jsearch_id == job_id).first()

        if not job:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Job not found",
            )

        # Check if already dismissed
        existing_action = db.query(UserJobAction).filter(
            UserJobAction.user_id == current_user.id,
            UserJobAction.job_id == job.id,
            UserJobAction.action == "dismissed",
        ).first()

        if existing_action:
            return {"message": "Job already dismissed", "action": "dismissed", "job_id": job.id}

        # Remove from saved if exists
        saved_action = db.query(UserJobAction).filter(
            UserJobAction.user_id == current_user.id,
            UserJobAction.job_id == job.id,
            UserJobAction.action == "saved",
        ).first()

        if saved_action:
            db.delete(saved_action)

        # Create dismiss action
        action = UserJobAction(
            user_id=current_user.id,
            job_id=job.id,
            action="dismissed",
        )
        db.add(action)
        db.commit()

        logger.info(f"User {current_user.id} dismissed job {job_id}")

        return {"message": "Job dismissed", "action": "dismissed", "job_id": job.id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error dismissing job: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to dismiss job",
        )


@router.post("/jobs/{job_id}/apply")
def apply_to_job(
    job_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Mark a job as applied and get apply link."""
    try:
        # Get job
        job = db.query(Job).filter(Job.jsearch_id == job_id).first()

        if not job:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Job not found",
            )

        # Check if already applied
        existing_action = db.query(UserJobAction).filter(
            UserJobAction.user_id == current_user.id,
            UserJobAction.job_id == job.id,
            UserJobAction.action == "applied",
        ).first()

        if existing_action:
            return {
                "message": "Already marked as applied",
                "action": "applied",
                "job_id": job.id,
                "apply_url": job.apply_url,
            }

        # Create apply action
        action = UserJobAction(
            user_id=current_user.id,
            job_id=job.id,
            action="applied",
        )
        db.add(action)

        # Remove from dismissed if was dismissed
        dismissed_action = db.query(UserJobAction).filter(
            UserJobAction.user_id == current_user.id,
            UserJobAction.job_id == job.id,
            UserJobAction.action == "dismissed",
        ).first()

        if dismissed_action:
            db.delete(dismissed_action)

        db.commit()

        logger.info(f"User {current_user.id} applied to job {job_id}")

        return {
            "message": "Job marked as applied",
            "action": "applied",
            "job_id": job.id,
            "apply_url": job.apply_url,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error applying to job: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to apply to job",
        )


@router.post("/search-async")
def search_jobs_async(
    request: JobSearchRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Start async job search. Returns search_id immediately."""
    try:
        # Create search record
        search_id = str(uuid.uuid4())
        search = JobSearch(
            id=search_id,
            user_id=current_user.id,
            status="pending",
            filters=json.dumps({
                "limit": request.limit,
                "required_skills": request.required_skills,
                "force_refresh": request.force_refresh,
            }),
        )
        db.add(search)
        db.commit()

        logger.info(f"Created async search {search_id} for user {current_user.id}")

        # Start background task
        start_background_search(
            search_id,
            str(current_user.id),
            {
                "limit": request.limit,
                "required_skills": request.required_skills,
                "force_refresh": request.force_refresh,
            },
        )

        return {
            "search_id": search_id,
            "status": "pending",
            "message": "Job search started. You will be notified when results are ready.",
        }

    except Exception as e:
        logger.error(f"Error starting async search: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to start job search",
        )


@router.get("/search-status/{search_id}")
def get_search_status(
    search_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get status of async job search."""
    try:
        search = db.query(JobSearch).filter(
            JobSearch.id == search_id,
            JobSearch.user_id == current_user.id,
        ).first()

        if not search:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Search not found",
            )

        return {
            "search_id": search_id,
            "status": search.status,
            "total_jobs_searched": search.total_jobs_searched,
            "total_matches": search.total_matches,
            "created_at": search.created_at,
            "completed_at": search.completed_at,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting search status: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to get search status",
        )


@router.get("/search-results/{search_id}")
def get_search_results(
    search_id: str,
    limit: int = 50,
    offset: int = 0,
    min_score: int = 60,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get results from async job search."""
    try:
        search = db.query(JobSearch).filter(
            JobSearch.id == search_id,
            JobSearch.user_id == current_user.id,
        ).first()

        if not search:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Search not found",
            )

        if search.status != "completed":
            return {
                "status": search.status,
                "matches": [],
                "total": 0,
            }

        # Get matches
        query = db.query(UserJobMatch).filter(
            UserJobMatch.user_id == current_user.id,
            UserJobMatch.match_score >= min_score,
        )

        total = query.count()

        matches = (
            query.order_by(UserJobMatch.match_score.desc())
            .offset(offset)
            .limit(limit)
            .all()
        )

        result_matches = []
        for match in matches:
            job = db.query(Job).filter(Job.id == match.job_id).first()
            if not job:
                continue

            action = db.query(UserJobAction).filter(
                UserJobAction.user_id == current_user.id,
                UserJobAction.job_id == match.job_id,
            ).first()

            match_reasons_data = match.match_reasons
            if isinstance(match_reasons_data, str):
                match_reasons_data = json.loads(match_reasons_data) if match_reasons_data else {}
            match_reasons_list = match_reasons_data.get("reasons", []) if match_reasons_data else []

            result_matches.append({
                "match_id": str(match.id),
                "job": {
                    "id": str(job.id),
                    "jsearch_id": job.jsearch_id,
                    "title": job.title,
                    "company": job.company,
                    "location": job.location,
                    "is_remote": job.is_remote,
                    "salary_min": job.salary_min,
                    "salary_max": job.salary_max,
                    "description": job.description,
                    "apply_url": job.apply_url,
                    "posted_at": job.posted_at,
                },
                "match_score": match.match_score,
                "match_reasons": match_reasons_list,
                "user_action": action.action if action else None,
            })

        return {
            "status": search.status,
            "total": total,
            "limit": limit,
            "offset": offset,
            "matches": result_matches,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting search results: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to get search results",
        )


@router.get("/search-results/{search_id}/export")
def export_search_results(
    search_id: str,
    min_score: int = 60,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export search results as Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        search = db.query(JobSearch).filter(
            JobSearch.id == search_id,
            JobSearch.user_id == current_user.id,
        ).first()

        if not search:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Search not found",
            )

        # Get all matches
        matches = db.query(UserJobMatch).filter(
            UserJobMatch.user_id == current_user.id,
            UserJobMatch.match_score >= min_score,
        ).order_by(UserJobMatch.match_score.desc()).all()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Matches"

        # Add headers
        headers = [
            "Rank",
            "Job Title",
            "Company",
            "Location",
            "Remote",
            "Match %",
            "Salary Min",
            "Salary Max",
            "Why Matched",
            "Status",
            "Apply URL",
        ]

        ws.append(headers)

        # Style header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data
        for rank, match in enumerate(matches, 1):
            job = db.query(Job).filter(Job.id == match.job_id).first()
            if not job:
                continue

            action = db.query(UserJobAction).filter(
                UserJobAction.user_id == current_user.id,
                UserJobAction.job_id == match.job_id,
            ).first()

            match_reasons_data = match.match_reasons
            if isinstance(match_reasons_data, str):
                match_reasons_data = json.loads(match_reasons_data) if match_reasons_data else {}
            match_reasons = match_reasons_data.get("reasons", []) if match_reasons_data else []

            ws.append([
                rank,
                job.title,
                job.company,
                job.location,
                "Yes" if job.is_remote else "No",
                f"{match.match_score}%",
                job.salary_min or "",
                job.salary_max or "",
                "; ".join(match_reasons[:3]) if match_reasons else "Good match",
                action.action.title() if action else "Not applied",
                job.apply_url,
            ])

        # Adjust column widths
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 40
        ws.column_dimensions["J"].width = 12
        ws.column_dimensions["K"].width = 40

        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        logger.info(f"Exported {len(matches)} matches for user {current_user.id}")

        return FileResponse(
            path=temp_file.name,
            filename=f"rolio_job_matches_{search_id[:8]}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel export requires openpyxl library",
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting results: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to export results",
        )
